import os
import datetime
import time
import sys
import openpyxl
import logging
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Font, colors, Alignment

def ExistSheet(sheetNames,asheetName):
    flage=False
    for asname in sheetNames:
        if(asname==asheetName):
            flage=True

    return flage

def getContenCell(mergedlist,acellpos):
    #for amerg in mergedlist:
        #print(amerg)
        #print(absolute_coordinate(amerg))
    return 'E2'



#拆分淀粉糖厂
def getFactory(orsheet,tarwb,tarsheetName):
    maxrow=orsheet.max_row
    if(maxrow<=0):
        return False
    #取ABCD四列
    orcol='d'+str(maxrow)
    cellrang=orsheet['A1':orcol]
    outsheet=tarwb.create_sheet(tarsheetName,0)
    for merg in orsheet.merged_cell_ranges:
        outsheet.merge_cells(merg)
    for arow in cellrang:
        for acell in arow:
            tarrow=acell.row
            tarcol=acell.col_idx
            if(acell.value==None):
                continue
            outsheet.cell(row=tarrow,column=tarcol).value=acell.value

    return True

#淀粉糖计算
def getSugar(orsheet,tarwb,tarsheetName):
    maxrow=orsheet.max_row
    maxcol=get_column_letter(orsheet.max_column)
    if(maxrow<=0):
        return False
    orcol=maxcol+str(maxrow)
    cellrang=orsheet['E4':orcol]
    outsheet=tarwb.create_sheet(tarsheetName,0)
    ormerged=orsheet.merged_cell_ranges
    outrowcount=1
    titleFlag=True
    title=['日期','糖类型','地区','厂家','日产量','日库存','日均价','时间差','年度','周','周度产量','周五库存','周均价']
    if(titleFlag):
        titleFlag=False
        outsheet.append(title)
    for arow in cellrang:
        for acell in arow:
            tarrow=acell.row
            tarcol=acell.col_idx
            #print(acell.value)
            if(acell.value==None):
                continue
            outrowcount+=1
            acol=column_index_from_string(acell.column)
            #日期
            #糖类型
            #地区
            #厂家
            #日产量
            atile=orsheet.cell(row=3,column=acol).value
            #日库存
            #日均价
            #时间差
            #年度
            ayear=orsheet.cell(row=1,column=acol).value
            #周
            #周度产量
            #周五库存
            #周均价

            aweek=orsheet.cell(row=2,column=acol).value
            print(ayear)
            #print(aweek)

    return True










start='20180101'
end='20180201'
sName_zhou='淀粉糖(周汇总)'
sName_factory='淀粉糖厂'

print('-------------欢迎使用淀粉糖文件截取工具----------------')
starttime =time.time()
while(len(start)==0 or len(end)==0):
    if(len(start)==0):
        start=input('请输入要截取淀粉糖文件的开始时间(如20180102）  \n开始：')
    if(len(end)==0):
        end=input('请输入要截取淀粉糖文件的结束时间(如20180102）    \n结束：')

startDT=datetime.datetime.strptime(start,'%Y%m%d')
endDT=datetime.datetime.strptime(end,'%Y%m%d')
startD=datetime.date(startDT.year,startDT.month,startDT.day)
endD=datetime.date(endDT.year,endDT.month,endDT.day)

disday=(endD-startD).days+1
if(disday<=0):
    input("日期输入有误请检查！！开始日期："+start+",结束日期"+end)
    sys.exit(0)



ROOTPATH=os.path.dirname(os.path.realpath(__file__))+"\\"
OUTPAHT=ROOTPATH+"output\\"
if( not os.path.exists(OUTPAHT)):
    os.mkdir(OUTPAHT)
print("\n工作目录为："+ROOTPATH)
count=0
listfile=os.listdir(ROOTPATH)
print('===================================================\n')

for afile in listfile:
    if(afile.endswith('xlsx',len(afile)-4,len(afile))
    or afile.endswith('xls',len(afile)-3,len(afile))):
        jangshuifile=afile.find('淀粉糖')
        if(jangshuifile<0):
            print(afile+"不是淀粉糖文件。")
            continue
        else:

            xlsfile=ROOTPATH+afile
            spfile=afile.split('.')
            savexlsfile=OUTPAHT+spfile[0]+"_"+start+"-"+end+"."+spfile[1]
            outfile=spfile[0]+"_"+start+"-"+end+"."+spfile[1]
            count+=1
            print('-----------------------------------------------\n')
            print('现在开始处理：'+afile)
            print('现在开始生成：'+outfile+"\n请耐心请待！\n")
            readwb = load_workbook(filename=xlsfile,data_only=True)
            outwb=openpyxl.Workbook()

            shtNams=readwb.sheetnames
            if(ExistSheet(shtNams,sName_zhou)):
                print("正在处理sheet："+sName_zhou)
                zhousheet = readwb.get_sheet_by_name(sName_zhou)
                factoryF=True
                #factoryF=getFactory(zhousheet,outwb,sName_factory)
                sugarF=getSugar(zhousheet,outwb,sName_zhou)

            else:

                 #factoryF=False
                 sugarF=False

        if(factoryF or  sugarF):
           outwb.save(savexlsfile)
        else:
           print(afile+"没有指定日期的数据！")

endtime = time.time()
costime=endtime - starttime
print("------------------------------------------------------")
print('程序运行时间为：{:.0f}m {:.0f}s'.format(costime/60, costime % 60))
if(count==0):
    input("没有要处理的excle,系统退出！")
    sys.exit(0)
else:
    print('===================================================\n')
    input("所有文件处理完毕！请检查！生成文件目录为：\n"+OUTPAHT)
    sys.exit(0)
